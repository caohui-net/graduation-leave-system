from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.utils import timezone
from django.db import transaction
from django.http import HttpResponse
from django.db.models import Prefetch
import logging
from drf_spectacular.utils import extend_schema, OpenApiParameter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

logger = logging.getLogger(__name__)

from .models import Approval, ApprovalDecision, ApprovalStep
from .serializers import ApprovalSerializer, ApprovalActionSerializer, ApprovalListSerializer, ApprovalListResponseSerializer
from .pagination import ApprovalLimitOffsetPagination
from .validators import approval_step_matches_application_status
from apps.applications.models import Application, ApplicationStatus
from apps.users.models import User, UserRole
from apps.notifications.services import notify_approval_decided
from schema import ErrorResponseSerializer
import uuid


def sanitize_excel_formula(value):
    """Sanitize text to prevent Excel formula injection."""
    if not value:
        return value
    value_str = str(value)
    if value_str and value_str[0] in ('=', '+', '-', '@'):
        return "'" + value_str
    return value_str


@extend_schema(
    operation_id='approvals_list',
    summary='获取审批列表',
    description='获取当前用户的待审批列表（辅导员或学工部）',
    parameters=[
        OpenApiParameter('decision', str, description='决策过滤：pending/approved/rejected/all（默认pending）'),
        OpenApiParameter('name', str, description='按姓名查询'),
        OpenApiParameter('student_id', str, description='按学号查询'),
        OpenApiParameter('building', str, description='按楼栋查询'),
        OpenApiParameter('limit', int, description='每页数量（默认20）'),
        OpenApiParameter('offset', int, description='偏移量（默认0）'),
    ],
    responses={
        200: ApprovalListResponseSerializer,
        403: ErrorResponseSerializer,
    },
    tags=['审批']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_approvals(request):
    user = request.user

    # 学生禁止访问
    if user.role == UserRole.STUDENT:
        return Response(
            {'error': {'code': 'FORBIDDEN', 'message': '学生不能访问审批列表'}},
            status=status.HTTP_403_FORBIDDEN
        )

    # 宿管员: 只看自己的dorm_manager审批
    if user.role == UserRole.DORM_MANAGER:
        queryset = Approval.objects.filter(
            approver=user,
            step=ApprovalStep.DORM_MANAGER
        ).select_related('application', 'application__student', 'approver')
        default_decision = 'pending'

    # 辅导员: 只看自己的counselor审批
    elif user.role == UserRole.COUNSELOR:
        queryset = Approval.objects.filter(
            approver=user,
            step=ApprovalStep.COUNSELOR
        ).select_related('application', 'application__student', 'approver')
        default_decision = 'pending'

    # 学工部: 查看所有审批（存档用）
    elif user.role == UserRole.DEAN:
        queryset = Approval.objects.all().select_related('application', 'application__student', 'approver')
        default_decision = None

    # 学工管理员: 查看所有审批（管理用）
    elif user.role == UserRole.ADMIN:
        queryset = Approval.objects.all().select_related('application', 'application__student', 'approver')
        default_decision = None

    else:
        return Response(
            {'error': {'code': 'FORBIDDEN', 'message': '无效的用户角色'}},
            status=status.HTTP_403_FORBIDDEN
        )

    # Decision filtering (default to pending for counselor/dorm_manager)
    decision_param = request.query_params.get('decision', default_decision)
    if decision_param and decision_param != 'all':
        queryset = queryset.filter(decision=decision_param)

    # Application type filtering
    app_type = request.query_params.get('application_type')
    if app_type:
        queryset = queryset.filter(application__application_type=app_type)

    # 查询过滤
    student_name = request.query_params.get('student_name')
    if student_name:
        queryset = queryset.filter(application__student_name__icontains=student_name)

    student_id = request.query_params.get('student_id')
    if student_id:
        queryset = queryset.filter(application__student__user_id__icontains=student_id)

    class_id = request.query_params.get('class_id')
    if class_id:
        queryset = queryset.filter(application__class_id__icontains=class_id)

    building = request.query_params.get('building')
    if building:
        queryset = queryset.filter(application__student__building__icontains=building)

    # 排序
    queryset = queryset.order_by('-created_at', '-approval_id')

    # 分页
    paginator = ApprovalLimitOffsetPagination()
    page = paginator.paginate_queryset(queryset, request)

    # 序列化
    serializer = ApprovalListSerializer(page, many=True)

    return paginator.get_paginated_response(serializer.data)


@extend_schema(
    operation_id='approvals_get',
    summary='获取审批详情',
    description='获取指定审批的详细信息',
    responses={
        200: ApprovalSerializer,
        403: ErrorResponseSerializer,
        404: ErrorResponseSerializer,
    },
    tags=['审批']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_approval(request, approval_id):
    try:
        approval = Approval.objects.select_related('application', 'approver').get(approval_id=approval_id)
    except Approval.DoesNotExist:
        return Response({'error': {'code': 'NOT_FOUND', 'message': '审批记录不存在'}},
                        status=status.HTTP_404_NOT_FOUND)

    user = request.user

    # Permission check: only the approver or dean/admin can view this approval
    if user.role in [UserRole.DEAN, UserRole.ADMIN] or approval.approver_id == user.user_id:
        return Response(ApprovalSerializer(approval).data)

    return Response({'error': {'code': 'FORBIDDEN', 'message': '无权限访问此资源'}},
                    status=status.HTTP_403_FORBIDDEN)


@extend_schema(
    operation_id='approvals_approve',
    summary='通过审批',
    description='审批人通过指定的审批申请',
    request=ApprovalActionSerializer,
    responses={
        200: ApprovalSerializer,
        400: ErrorResponseSerializer,
        403: ErrorResponseSerializer,
        404: ErrorResponseSerializer,
        409: ErrorResponseSerializer,
    },
    tags=['审批']
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def approve_approval(request, approval_id):
    try:
        approval = Approval.objects.select_for_update().get(approval_id=approval_id)
    except Approval.DoesNotExist:
        return Response({'error': {'code': 'NOT_FOUND', 'message': '审批记录不存在'}},
                        status=status.HTTP_404_NOT_FOUND)

    user = request.user
    # Admin can approve/reject any step
    if user.role != UserRole.ADMIN:
        if approval.step == ApprovalStep.DORM_MANAGER and user.role != UserRole.DORM_MANAGER:
            return Response({'error': {'code': 'FORBIDDEN', 'message': '无权限执行此操作'}},
                            status=status.HTTP_403_FORBIDDEN)
        if approval.step == ApprovalStep.COUNSELOR and user.role != UserRole.COUNSELOR:
            return Response({'error': {'code': 'FORBIDDEN', 'message': '无权限执行此操作'}},
                            status=status.HTTP_403_FORBIDDEN)
        if approval.step == ApprovalStep.DEAN and user.role != UserRole.DEAN:
            return Response({'error': {'code': 'FORBIDDEN', 'message': '无权限执行此操作'}},
                            status=status.HTTP_403_FORBIDDEN)

        if approval.approver_id != user.user_id:
            return Response({'error': {'code': 'FORBIDDEN', 'message': '无权限执行此操作'}},
                            status=status.HTTP_403_FORBIDDEN)

    if approval.decision != ApprovalDecision.PENDING:
        return Response({'error': {'code': 'CONFLICT', 'message': '审批已完成，不能重复操作'}},
                        status=status.HTTP_409_CONFLICT)

    if not approval_step_matches_application_status(approval):
        return Response({'error': {'code': 'CONFLICT', 'message': '申请状态与审批步骤不匹配'}},
                        status=status.HTTP_409_CONFLICT)

    serializer = ApprovalActionSerializer(data=request.data)
    if not serializer.is_valid():
        return Response({'error': {'code': 'VALIDATION_ERROR', 'message': '请求参数验证失败'}},
                        status=status.HTTP_400_BAD_REQUEST)

    approval.decision = ApprovalDecision.APPROVED
    approval.comment = serializer.validated_data.get('comment', '')
    approval.decided_at = timezone.now()
    approval.decided_by = user
    approval.save()

    notify_approval_decided(approval)

    application = approval.application
    if approval.step == ApprovalStep.DORM_MANAGER:
        # Auto-complete other pending dorm manager approvals for the same building
        # (New requirement: any dorm manager in the building can approve, others see "already approved")
        other_dorm_approvals = Approval.objects.filter(
            application=application,
            step=ApprovalStep.DORM_MANAGER,
            decision=ApprovalDecision.PENDING
        ).exclude(approval_id=approval.approval_id)

        if other_dorm_approvals.exists():
            now = timezone.now()
            actual_approver = approval.decided_by or approval.approver
            count = other_dorm_approvals.count()
            for other_approval in other_dorm_approvals:
                other_approval.decision = ApprovalDecision.APPROVED
                other_approval.comment = f'已由{approval.approver_name}完成审批，无需重复操作'
                other_approval.decided_at = now
                other_approval.decided_by = actual_approver
                other_approval.save()
            logging.info(
                f"Auto-completed {count} other dorm manager approvals "
                f"for application {application.application_id} after approval by {actual_approver.user_id}"
            )

        # Get counselor by department (Phase 3 design: department-based routing)
        # Note: Original design used ClassMapping (class_id), but Phase 3 user requirements
        # changed to "按学院向辅导员审批" (approval by department/college).
        # Multiple counselors per department are allowed (different classes within department).
        # Selection: order_by('user_id') picks lowest ID for deterministic routing.
        counselors = User.objects.filter(
            role=UserRole.COUNSELOR,
            department=application.student.department,
            active=True
        ).order_by('user_id')

        if counselors.count() > 1:
            logging.warning(
                f"Multiple counselors found for department {application.student.department}: "
                f"{counselors.count()} matches. Selected {counselors.first().user_id} via order_by('user_id')"
            )

        counselor = counselors.first()

        if not counselor:
            return Response({'error': {'code': 'NOT_FOUND', 'message': '该学院辅导员不存在',
                                        'details': {'department': application.student.department}}},
                            status=status.HTTP_404_NOT_FOUND)

        # Atomic counselor approval creation (防止竞态条件)
        counselor_approval, created = Approval.objects.get_or_create(
            application=application,
            step=ApprovalStep.COUNSELOR,
            defaults={
                'approval_id': f'apv_{uuid.uuid4().hex[:8]}',
                'approver': counselor,
                'approver_name': counselor.name,
                'decision': ApprovalDecision.PENDING
            }
        )

        if not created:
            return Response({'error': {'code': 'CONFLICT', 'message': '辅导员审批已存在，不能重复创建'}},
                            status=status.HTTP_409_CONFLICT)

        application.status = ApplicationStatus.PENDING_COUNSELOR
        application.save()
    elif approval.step == ApprovalStep.COUNSELOR:
        # Counselor approval completes the process (2-level approval)
        application.status = ApplicationStatus.APPROVED
        application.save()

    return Response(ApprovalSerializer(approval).data)


@extend_schema(
    operation_id='approvals_reject',
    summary='驳回审批',
    description='审批人驳回指定的审批申请',
    request=ApprovalActionSerializer,
    responses={
        200: ApprovalSerializer,
        400: ErrorResponseSerializer,
        403: ErrorResponseSerializer,
        404: ErrorResponseSerializer,
        409: ErrorResponseSerializer,
    },
    tags=['审批']
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def reject_approval(request, approval_id):
    try:
        approval = Approval.objects.select_for_update().get(approval_id=approval_id)
    except Approval.DoesNotExist:
        return Response({'error': {'code': 'NOT_FOUND', 'message': '审批记录不存在'}},
                        status=status.HTTP_404_NOT_FOUND)

    user = request.user
    # Admin can approve/reject any step
    if user.role != UserRole.ADMIN:
        if approval.step == ApprovalStep.DORM_MANAGER and user.role != UserRole.DORM_MANAGER:
            return Response({'error': {'code': 'FORBIDDEN', 'message': '无权限执行此操作'}},
                            status=status.HTTP_403_FORBIDDEN)
        if approval.step == ApprovalStep.COUNSELOR and user.role != UserRole.COUNSELOR:
            return Response({'error': {'code': 'FORBIDDEN', 'message': '无权限执行此操作'}},
                            status=status.HTTP_403_FORBIDDEN)
        if approval.step == ApprovalStep.DEAN and user.role != UserRole.DEAN:
            return Response({'error': {'code': 'FORBIDDEN', 'message': '无权限执行此操作'}},
                            status=status.HTTP_403_FORBIDDEN)

        if approval.approver_id != user.user_id:
            return Response({'error': {'code': 'FORBIDDEN', 'message': '无权限执行此操作'}},
                            status=status.HTTP_403_FORBIDDEN)

    if approval.decision != ApprovalDecision.PENDING:
        return Response({'error': {'code': 'CONFLICT', 'message': '审批已完成，不能重复操作'}},
                        status=status.HTTP_409_CONFLICT)

    if not approval_step_matches_application_status(approval):
        return Response({'error': {'code': 'CONFLICT', 'message': '申请状态与审批步骤不匹配'}},
                        status=status.HTTP_409_CONFLICT)

    serializer = ApprovalActionSerializer(data=request.data)
    if not serializer.is_valid():
        return Response({'error': {'code': 'VALIDATION_ERROR', 'message': '请求参数验证失败'}},
                        status=status.HTTP_400_BAD_REQUEST)

    approval.decision = ApprovalDecision.REJECTED
    approval.comment = serializer.validated_data.get('comment', '')
    approval.decided_at = timezone.now()
    approval.decided_by = user
    approval.save()

    notify_approval_decided(approval)

    application = approval.application
    application.status = ApplicationStatus.REJECTED
    application.save()

    return Response(ApprovalSerializer(approval).data)


@extend_schema(
    operation_id='approvals_export',
    summary='导出审批数据',
    description='导出所有审批数据到Excel（仅学工部）',
    responses={
        200: {'description': 'Excel文件'},
        403: ErrorResponseSerializer,
    },
    tags=['审批']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_approvals(request):
    try:
        user = request.user

        if user.role not in [UserRole.DEAN, UserRole.ADMIN, UserRole.COUNSELOR, UserRole.DORM_MANAGER]:
            return Response(
                {'error': {'code': 'FORBIDDEN', 'message': '仅管理人员可导出数据'}},
                status=status.HTTP_403_FORBIDDEN
            )

        # Export with permission filtering (same logic as list_approvals)
        from django.db.models import OuterRef, Subquery, Prefetch

        # Get application_type filter
        app_type = request.query_params.get('application_type', 'leave_school')

        # Apply role-based filtering
        if user.role == UserRole.DORM_MANAGER:
            # Only export students from own approval list
            applications = Application.objects.filter(
                application_type=app_type,
                approvals__approver=user,
                approvals__step=ApprovalStep.DORM_MANAGER
            ).distinct().select_related('student').order_by('student__user_id')

        elif user.role == UserRole.COUNSELOR:
            # Only export students from own approval list
            applications = Application.objects.filter(
                application_type=app_type,
                approvals__approver=user,
                approvals__step=ApprovalStep.COUNSELOR
            ).distinct().select_related('student').order_by('student__user_id')

        else:  # DEAN or ADMIN
            # Export all applications of this type
            applications = Application.objects.filter(
                application_type=app_type
            ).select_related('student').order_by('student', '-created_at').distinct('student')

        # For role-specific queries, also get latest per student
        if user.role in [UserRole.DORM_MANAGER, UserRole.COUNSELOR]:
            applications = applications.order_by('student', '-created_at').distinct('student')

        # Fetch with prefetch for efficiency
        application_ids = list(applications.values_list('application_id', flat=True))
        applications = Application.objects.filter(
            application_id__in=application_ids
        ).prefetch_related(
            Prefetch('approvals', queryset=Approval.objects.filter(step=ApprovalStep.DORM_MANAGER), to_attr='dorm_approvals_list'),
            Prefetch('approvals', queryset=Approval.objects.filter(step=ApprovalStep.COUNSELOR), to_attr='counselor_approvals_list')
        ).select_related('student').order_by('student__user_id')

        wb = Workbook()
        ws = wb.active
        ws.title = '学生数据'

        headers = ['提交人', '学号', '手机号', '离校日期', '楼栋号', '房间号', '提交时间', '审批状态',
                   '宿管员', '宿管审批时间', '宿管审批结果',
                   '辅导员', '辅导员审批时间', '辅导员审批结果']
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for app in applications:
            dorm_approval = app.dorm_approvals_list[0] if app.dorm_approvals_list else None
            counselor_approval = app.counselor_approvals_list[0] if app.counselor_approvals_list else None

            status_display = app.get_status_display()
            leave_date = app.leave_date.strftime('%Y-%m-%d') if app.leave_date else ''
            contact_phone = app.contact_phone or app.student.phone or ''
            submit_time = app.created_at.strftime('%Y-%m-%d %H:%M:%S')

            row = [
                sanitize_excel_formula(app.student.name),
                sanitize_excel_formula(app.student.user_id),
                sanitize_excel_formula(contact_phone),
                leave_date,
                sanitize_excel_formula(app.student.building or ''),
                sanitize_excel_formula(app.student.room_number or ''),
                submit_time,
                status_display,
                sanitize_excel_formula(dorm_approval.approver_name if dorm_approval else ''),
                dorm_approval.decided_at.strftime('%Y-%m-%d %H:%M:%S') if dorm_approval and dorm_approval.decided_at else '',
                dorm_approval.get_decision_display() if dorm_approval else '',
                sanitize_excel_formula(counselor_approval.approver_name if counselor_approval else ''),
                counselor_approval.decided_at.strftime('%Y-%m-%d %H:%M:%S') if counselor_approval and counselor_approval.decided_at else '',
                counselor_approval.get_decision_display() if counselor_approval else '',
            ]
            ws.append(row)

        # Set fixed column widths (faster than auto-sizing for large datasets)
        column_widths = [12, 12, 15, 12, 10, 10, 20, 12, 12, 20, 12, 12, 20, 12]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename_prefix = '留校申请' if app_type == 'stay_school' else '离校申请'
        response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)

        return response
    except Exception as e:
        logging.error(f'Export failed: {type(e).__name__}: {e}', exc_info=True)
        return Response(
            {'error': {'code': 'INTERNAL_ERROR', 'message': f'导出失败: {str(e)}'}},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@extend_schema(
    operation_id='approvals_statistics',
    summary='获取审批统计',
    description='获取当前用户的审批统计数据（总提交数、待审批数、已通过数）',
    responses={
        200: {
            'description': '统计数据',
            'content': {
                'application/json': {
                    'example': {
                        'total': 100,
                        'pending': 10,
                        'approved': 85,
                        'rejected': 5
                    }
                }
            }
        },
        403: ErrorResponseSerializer,
    },
    tags=['审批']
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_statistics(request):
    user = request.user

    # 学生禁止访问
    if user.role == UserRole.STUDENT:
        return Response(
            {'error': {'code': 'FORBIDDEN', 'message': '学生不能访问统计数据'}},
            status=status.HTTP_403_FORBIDDEN
        )

    # 根据角色筛选数据
    if user.role == UserRole.DORM_MANAGER:
        queryset = Approval.objects.filter(
            approver=user,
            step=ApprovalStep.DORM_MANAGER
        )
    elif user.role == UserRole.COUNSELOR:
        queryset = Approval.objects.filter(
            approver=user,
            step=ApprovalStep.COUNSELOR
        )
    elif user.role in [UserRole.DEAN, UserRole.ADMIN]:
        queryset = Approval.objects.all()
    else:
        return Response(
            {'error': {'code': 'FORBIDDEN', 'message': '无效的用户角色'}},
            status=status.HTTP_403_FORBIDDEN
        )

    # Filter by application_type
    app_type = request.query_params.get('application_type')
    if app_type:
        queryset = queryset.filter(application__application_type=app_type)

    # 统计数据（按去重的学生计算，避免同一学生多次申请被重复统计）
    total_students = queryset.values('application__student').distinct().count()
    pending = queryset.filter(decision=ApprovalDecision.PENDING).values('application__student').distinct().count()
    approved = queryset.filter(decision=ApprovalDecision.APPROVED).values('application__student').distinct().count()
    rejected = queryset.filter(decision=ApprovalDecision.REJECTED).values('application__student').distinct().count()

    return Response({
        'total': total_students,
        'pending': pending,
        'approved': approved,
        'rejected': rejected
    })


@extend_schema(
    operation_id='approvals_batch_action',
    summary='批量处理审批',
    request={
        'application/json': {
            'type': 'object',
            'properties': {
                'approval_ids': {'type': 'array', 'items': {'type': 'string'}},
                'action': {'type': 'string', 'enum': ['approve', 'reject']},
                'comment': {'type': 'string'}
            },
            'required': ['approval_ids', 'action']
        }
    },
    responses={200: {'description': '批量处理成功'}, 400: ErrorResponseSerializer, 403: ErrorResponseSerializer},
    tags=['审批']
)
@api_view(['POST'])
@permission_classes([IsAuthenticated])
@transaction.atomic
def batch_action_approvals(request):
    approval_ids = request.data.get('approval_ids', [])
    action = request.data.get('action')
    comment = request.data.get('comment', '')

    if not approval_ids or not isinstance(approval_ids, list):
        return Response({'error': {'code': 'VALIDATION_ERROR', 'message': '审批ID列表不能为空'}}, status=status.HTTP_400_BAD_REQUEST)
    if action not in ['approve', 'reject']:
        return Response({'error': {'code': 'VALIDATION_ERROR', 'message': 'action必须为approve或reject'}}, status=status.HTTP_400_BAD_REQUEST)

    user = request.user
    approvals = Approval.objects.select_for_update().filter(approval_id__in=approval_ids, approver=user, decision=ApprovalDecision.PENDING)

    if approvals.count() != len(approval_ids):
        missing_count = len(approval_ids) - approvals.count()
        found_ids = set(approvals.values_list('approval_id', flat=True))
        missing_ids = [aid for aid in approval_ids if aid not in found_ids]

        # Log detailed error for debugging
        logger.error(f"Batch action rejected for user {user.user_id}: {missing_count} approvals not accessible. Missing IDs: {missing_ids[:10]}")

        # Check reason for first missing approval
        if missing_ids:
            sample_approval = Approval.objects.filter(approval_id=missing_ids[0]).first()
            if sample_approval:
                logger.error(f"Sample missing approval {missing_ids[0]}: approver={sample_approval.approver_id}, decision={sample_approval.decision}, step={sample_approval.step}")
            else:
                logger.error(f"Sample missing approval {missing_ids[0]}: not found in database")

        return Response({
            'error': {
                'code': 'VALIDATION_ERROR',
                'message': f'无法操作 {missing_count} 个审批：不存在、非待审批状态或您无权限（只能操作分配给您的审批）',
                'details': {
                    'missing_count': missing_count,
                    'missing_ids': missing_ids[:10],  # 最多返回10个ID
                    'total_requested': len(approval_ids),
                    'valid_count': approvals.count()
                }
            }
        }, status=status.HTTP_400_BAD_REQUEST)

    now = timezone.now()
    decision = ApprovalDecision.APPROVED if action == 'approve' else ApprovalDecision.REJECTED

    for approval in approvals:
        if not approval_step_matches_application_status(approval):
            return Response({'error': {'code': 'CONFLICT', 'message': f'审批状态不匹配'}}, status=status.HTTP_409_CONFLICT)

        approval.decision = decision
        approval.comment = comment
        approval.decided_at = now
        approval.decided_by = user
        approval.save()

        notify_approval_decided(approval)

        application = approval.application
        if decision == ApprovalDecision.APPROVED:
            if approval.step == ApprovalStep.DORM_MANAGER:
                # Auto-complete other pending dorm manager approvals
                other_dorm_approvals = Approval.objects.filter(
                    application=application,
                    step=ApprovalStep.DORM_MANAGER,
                    decision=ApprovalDecision.PENDING
                ).exclude(approval_id=approval.approval_id)

                if other_dorm_approvals.exists():
                    for other_approval in other_dorm_approvals:
                        other_approval.decision = ApprovalDecision.APPROVED
                        other_approval.comment = f'已由{approval.approver_name}完成审批，无需重复操作'
                        other_approval.decided_at = now
                        other_approval.decided_by = user
                        other_approval.save()

                # Create counselor approval if not exists
                if not Approval.objects.filter(application=application, step=ApprovalStep.COUNSELOR).exists():
                    counselors = User.objects.filter(
                        role=UserRole.COUNSELOR,
                        department=application.student.department,
                        active=True
                    ).order_by('user_id')

                    counselor = counselors.first()
                    if not counselor:
                        return Response({'error': {'code': 'NOT_FOUND', 'message': '该学院辅导员不存在',
                                                    'details': {'department': application.student.department}}},
                                        status=status.HTTP_404_NOT_FOUND)

                    Approval.objects.create(
                        approval_id=f'apv_{uuid.uuid4().hex[:8]}',
                        application=application,
                        step=ApprovalStep.COUNSELOR,
                        approver=counselor,
                        approver_name=counselor.name,
                        decision=ApprovalDecision.PENDING
                    )

                application.status = ApplicationStatus.PENDING_COUNSELOR
            else:
                application.status = ApplicationStatus.APPROVED
        else:
            application.status = ApplicationStatus.REJECTED
        application.save()

    return Response({'success': True, 'processed': approvals.count()})
